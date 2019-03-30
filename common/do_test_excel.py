from openpyxl import load_workbook
from common.proje_path import cases_path
from common.test_config import TestConfig

class DoExcel:


    def data_read(self,sheet_name):
        '''读取用例，将每条用例分别存储在字典内，然后将所有字典存在列表
        :sheet_name  表单名'''
        wb=load_workbook(cases_path)
        #打开表格
        sheet=wb[sheet_name]
        #定位表单，self.sheet---参数化表单名
        data_list=[]
        for i in range(2,sheet.max_row+1):
            data_dict={}
            data_dict['用例序号']=sheet.cell(i,1).value
            data_dict['模块']=sheet.cell(i,2).value
            data_dict['请求方式']=sheet.cell(i,3).value
            data_dict['接口地址']=sheet.cell(i,4).value
            data_dict['用例名']=sheet.cell(i,5).value
            if sheet.cell(i,6).value.find("TEL")!=-1:
                #判断参数列是否含有字段"TEL"
                res=sheet.cell(i,6).value.replace("TEL",str(self.read_tel()))
                #存在"TEL"字段，则将该字段替换为表单TEL中的电话号码
                self.change_tel()
                #每次替换完"TEL"字段后，就将号码+1
                data_dict["参数"]=res
                #将修改完的参数列添加到字典中
            else:
                data_dict['参数']=sheet.cell(i,6).value
            data_dict['sql']=sheet.cell(i,7).value
            data_dict['预期结果']=sheet.cell(i,8).value
            #将用例以键值对形式存储在字典内
            data_list.append(data_dict)
            #将用例添加到列表

        final_cases=[]
        res=TestConfig().get_data('CASE','cases')
        #读取配置文件
        if eval(res)=='all':
            #判断读取到的内容，控制用例读取的条数，从配置文件获取到的内容默认为字符串
            final_cases=data_list
        else:
            for i in eval(res):
                final_cases.append(data_list[i-1])

        return final_cases
        #返回最终读取到的用例
    def read_tel(self):
        '''读取电话号码'''
        wb=load_workbook(cases_path)
        sheet=wb['TEL']
        tel=sheet.cell(1,2).value
        return tel
    def change_tel(self):
        '''修改电话号码，每次读取号码最后后一位+1'''
        wb=load_workbook(cases_path)
        sheet=wb['TEL']
        sheet.cell(1,2).value=self.read_tel()+1
        wb.save(cases_path)
        wb.close()

    def data_write(self,sheet_name,row,col,value):
        wb=load_workbook(cases_path)
        sheet=wb[sheet_name]
        sheet.cell(row,col).value=value
        wb.save(cases_path)
        wb.close()

if __name__ == '__main__':
    res=DoExcel().data_read('register')
    print(res)









